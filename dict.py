from collections import defaultdict
import openpyxl
import pickle

# Read the XLSX file into a dictionary
mydict = defaultdict(list)
wb = openpyxl.load_workbook('TRACKING_JULY_2024.xlsx')
ws = wb.active

# Create a defaultdict
mydict = defaultdict(list)

# Iterate over the rows in the worksheet
for row in ws.iter_rows(min_row=2, values_only=True):
    # Use the "Region name" column as the category
    mycategory = row[3] 
    # Create a dictionary from the row values and column names
    row_dict = dict(zip([cell.value for cell in ws[1]], row))
    row_dict.pop("REGION NAME", None)
    # Append the row dictionary to the category in the dictionary
    mydict[mycategory].append(row_dict)

# Convert back to a normal dictionary (optional)
mydict = dict(mydict)
with open('my_dict.pkl', 'wb') as f:
    pickle.dump(mydict, f)

